import os
import csv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# データを貼り付けるシートの名前
target_sheet_name = "1_台風データ貼り付け"

# 出力フォルダ
output_folder = "output"

# 出力フォルダ内の全ての数字6桁のフォルダを処理
for folder_name in os.listdir(output_folder):
    if len(folder_name) == 6 and folder_name.isdigit():
        folder_path = os.path.join(output_folder, folder_name)
        
        # CSVファイルの読み込み
        csv_files = [file for file in os.listdir(folder_path) if file.endswith('.csv')]
        if len(csv_files) == 1:  # CSVファイルが1つだけ存在する場合
            csv_filename = csv_files[0]
            csv_filepath = os.path.join(folder_path, csv_filename)
            
            # 対応するExcelファイルを検索
            xlsx_files = [file for file in os.listdir(folder_path) if file.endswith('.xlsx')]
            for xlsx_filename in xlsx_files:
                xlsx_filepath = os.path.join(folder_path, xlsx_filename)
                
                # Excelファイルをロード
                wb = load_workbook(xlsx_filepath)
                if target_sheet_name in wb.sheetnames:
                    ws = wb[target_sheet_name]
                    
                    # CSVファイルの内容をExcelに貼り付け
                    with open(csv_filepath, 'r', encoding='shift-jis') as csvfile:
                        csvreader = csv.reader(csvfile)
                        for row_index, row in enumerate(csvreader, start=1):
                            for col_index, value in enumerate(row, start=1):
                                try:
                                    value = float(value)  # 数値に変換
                                except ValueError:
                                    pass  # 変換できない場合はそのまま
                                ws.cell(row=row_index, column=col_index, value=value)
                    
                    # Excelファイルを保存
                    wb.save(xlsx_filepath)
                    print(f"Data from '{csv_filename}' has been pasted into '{xlsx_filename}'.")
                    break  # 1つのExcelファイルについて処理を完了したら終了
            else:
                print(f"No corresponding Excel file found in '{folder_name}'.")
        else:
            print(f"No CSV file or multiple CSV files found in '{folder_name}'.")
