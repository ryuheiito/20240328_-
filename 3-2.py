import os
import pandas as pd
from openpyxl import load_workbook

# フォルダのパス
csv_folder = 'output/out'
top_folder = 'output'

# CSVファイルを読み取ってxlsxファイルに貼り付ける関数
def paste_data_to_excel(csv_file, xlsx_file):
    # CSVファイルからデータを読み取る
    csv_data = pd.read_csv(csv_file, header=None, skiprows=5)
    # xlsxファイルを読み込む
    wb = load_workbook(filename=xlsx_file)
    # sheetを選択
    ws = wb['3_距離貼り付け']
    # 6行目からデータを貼り付ける
    for i, row in enumerate(csv_data.values, start=6):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j).value = value
    # xlsxファイルを保存
    wb.save(xlsx_file)

# 指定されたフォルダから再帰的にxlsxファイルを探す関数
def find_xlsx_files(folder, csv_filename):
    for root, dirs, files in os.walk(folder):
        for file in files:
            if file.endswith('.xlsx'):
                xlsx_file = os.path.join(root, file)
                # xlsxファイルの名前がcsvファイルと一致するかを確認
                if os.path.splitext(os.path.basename(xlsx_file))[0] == os.path.splitext(csv_filename)[0]:
                    return xlsx_file
    return None

# CSVフォルダ内のファイルを処理
for csv_file in os.listdir(csv_folder):
    if csv_file.endswith('.csv'):
        csv_path = os.path.join(csv_folder, csv_file)
        # 対応するxlsxファイルを探す
        xlsx_file = find_xlsx_files(top_folder, csv_file)
        if xlsx_file:
            paste_data_to_excel(csv_path, xlsx_file)
        else:
            print(f"対応するxlsxファイルが見つかりませんでした: {csv_file}")
